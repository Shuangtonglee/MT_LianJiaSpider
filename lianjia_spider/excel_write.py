from openpyxl import Workbook,load_workbook
from xpinyin import Pinyin
from time import time
from lianjia_spider.data_get import XiaoQu,Chengjiao

xiaoqu = XiaoQu()
chengjiao = Chengjiao()
ts = time()

def toNumber(string):
    try:
        return eval(string)
    except:
        return ''



class ExcelWrite:

    def __init__(self):
        pass

    def xiaoqu_write_into_excel(self):
        regions = ['浦东','闵行','宝山','徐汇','普陀','杨浦','长宁','松江','嘉定','黄浦',
          '静安','闸北','虹口','青浦','奉贤','金山','崇明','上海周边']
        #regions = ['闵行','虹口','浦东'] #测试
        p = Pinyin()
        filename = '上海链家小区数据.xlsx'
        sheet_number = 0
        wb = Workbook()
        ws = []
        for region_name in regions:
            print(region_name)
            ws.append(wb.create_sheet(region_name))
            if region_name  == '闵行':
                region_name  = 'minhang'
            else:
                region_name  = p.get_pinyin(region_name ,'')   #汉字转为拼音以符合链接要求
            region_name = p.get_pinyin(region_name,'')
            urls = xiaoqu.url(region_name)
            xiaoqu_list = xiaoqu.xiaoqu_data(urls)
            ws[sheet_number].append(['小区名称','挂牌均价（元/平）','在售数量(套)','区域','位置','地铁','建造时间'])
            for xq in xiaoqu_list:
                print(xq)
                ws[sheet_number].append([xq[0],toNumber(xq[1]),toNumber(xq[2]),xq[3],xq[4],xq[5],xq[6]])
            sheet_number += 1
        wb.save(filename=filename)
        print('Took {}s'.format(time() - ts))





    def chengjiao_write_into_excel(self):
        filename = '上海链家各区域小区二手房成交数据.xlsx'
        sheet_number = 0
        wb = Workbook()
        ws = []

        wb0 = load_workbook(filename='上海链家小区数据.xlsx')       #载入excel 表格
        region_names = wb0.get_sheet_names()                       #获取各sheet 名称，即各区名称
        for region_name in region_names[1:]:
            print(region_name)
            ws.append(wb.create_sheet(region_name))
            ws[sheet_number].append(['小区名称','户型','面积(平米)','楼层','成交单价（元/平）','成交总价(万)','挂牌总价(万)','地铁','朝向','装修','房本','电梯','房屋类型','成交周期(天)','成交时间','链接'])
            ws0 = wb0.get_sheet_by_name(region_name)            #根据sheet名称激活sheet
            for rx in range(2,ws0.max_row+1):
            #for rx in range(2,3):
                xiaoqu_name = ws0.cell(row=rx,column=1).value   #各sheet 第一列小区名称
                print(xiaoqu_name)
                urls = chengjiao.url(xiaoqu_name)
                if urls == []: #小区成交数量为0,循环直接跳到下一个小区
                    continue
                xiaoqu_chengjiao_list = chengjiao.chengjiao_data(urls)
                for xiaoqu_chengjiao_info in xiaoqu_chengjiao_list:
                    print(xiaoqu_chengjiao_info)
                    ws[sheet_number].append([xiaoqu_chengjiao_info[0],xiaoqu_chengjiao_info[1],toNumber(xiaoqu_chengjiao_info[2]),xiaoqu_chengjiao_info[3],toNumber(xiaoqu_chengjiao_info[4]),toNumber(xiaoqu_chengjiao_info[5]),
                                               toNumber(xiaoqu_chengjiao_info[6]),xiaoqu_chengjiao_info[7],xiaoqu_chengjiao_info[8],xiaoqu_chengjiao_info[9],xiaoqu_chengjiao_info[10],xiaoqu_chengjiao_info[11],xiaoqu_chengjiao_info[12],toNumber(xiaoqu_chengjiao_info[13]),xiaoqu_chengjiao_info[14],xiaoqu_chengjiao_info[15]])
            sheet_number +=1
        wb.save(filename=filename)
        print('Took {}s'.format(time() - ts))


